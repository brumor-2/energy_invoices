import pdfplumber  
import re
from collections import defaultdict
import os
from openpyxl.styles import PatternFill

lista_escolas = {
    #Número da instalação : ("Nome da escola relacionado ao número de instalação", nº da escola)
    # Em um exemplo prático, fica assim:
    # "104227": ("PROF. HENRIQUE GONÇALVES", 1),
    
}



def extract_school_name(pdf_path, lista_escolas):
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text()
        
        for school in lista_escolas:
            if school in text:
                return school
    return None


import itertools


def extract_numero_instalacao(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]
        
        # Defina as coordenadas da caixa delimitadora (x0, top, x1, bottom)
        # para a área onde o Número de Instalação deve estar localizado
        bbox = (0, 0, 300, 100)
        
        cropped_page = first_page.within_bbox(bbox)
        text = cropped_page.extract_text()
        
        match = re.search(r'\d+', text)
        
        if match:
            return match.group(0)
        else:
            print(f"Número de Instalação não encontrado no arquivo: {pdf_path}")
            return None






def get_sorted_schools_and_files(directory, lista_escolas):
    schools_and_files = defaultdict(list)
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                filepath = os.path.join(root, file)
                school_name = extract_school_name(filepath, lista_escolas)
                if school_name:
                    schools_and_files[school_name].append(filepath)
    sorted_schools = dict(sorted(schools_and_files.items(), key=lambda x: lista_escolas[x[0]]))
    return sorted_schools

directory = "C:\\path"
sorted_schools = get_sorted_schools_and_files(directory, lista_escolas)

for school in lista_escolas:
    if school not in sorted_schools:
        sorted_schools[school] = []

for school_number, filepaths in sorted_schools.items():
    school_name = lista_escolas[school_number][0]
    for filepath in filepaths:
        numero_instalacao = extract_numero_instalacao(filepath)
        if numero_instalacao:
            # Adicionar esta linha para criar e definir a variável 'renamed_filepath'
            renamed_filepath = os.path.join(os.path.dirname(filepath), f"{lista_escolas[school_number][1]}-ENEL-{school_name}-{numero_instalacao} - Venc Mai-23.pdf")
            
            # Verificar se o arquivo renomeado já existe antes de tentar renomeá-lo
            if not os.path.exists(renamed_filepath):
                os.rename(filepath, renamed_filepath)  # Adicione esta linha para renomear o arquivo
                print(f"Arquivo renomeado: {filepath} -> {renamed_filepath}")
            else:
                print(f"Arquivo já existe, ignorando: {renamed_filepath}")
        else:
            print(f"Não foi possível renomear o arquivo: {filepath}")



        

#Esta função corrigirá o erro do sistema mostrar números com . e . ao invés de . e ,
def sanitize_number_string(number_string):
    number_string = number_string.strip()
    sanitized_string = number_string.replace(".", "").replace(",", ".")
    return sanitized_string


# valor das faturas


def extract_valor(pdf_path):
    valor_pattern = r"((?:\d{1,3}\.?)+,\d{2})"
    
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]
        
        # Defina as coordenadas da caixa delimitadora (x0, top, x1, bottom)
        # para a área onde o valor da fatura deve estar localizado
        # Estes valores podem ser ajustados de acordo com os seus documentos
        bbox = (350, 50, 550, 150)
        
        cropped_page = first_page.within_bbox(bbox)
        text = cropped_page.extract_text()

        match = re.search(valor_pattern, text)
        
        if match:
            valor = match.group(1)
            valor = sanitize_number_string(valor)
            return float(valor)

        else:
            print(f"Valor não encontrado no arquivo: {pdf_path}")
            return None

def extract_numerodaconta(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]
        
        # Defina as coordenadas da caixa delimitadora (x0, top, x1, bottom)
        # para a área onde o Código de Barras deve estar localizado
        bbox = (0.0, 0.0, 595.276, 841.89) # Ajuste esses valores de acordo com a posição do código de barras em seus documentos
        
        cropped_page = first_page.within_bbox(bbox)
        text = cropped_page.extract_text()
        
        match = re.search(r'\d{12,14}', text)
        
        if match:
            return match.group(0)
        else:
            print(f"Número da Conta não encontrada no arquivo: {pdf_path}")
            return None
        



#criar planilha:

import openpyxl

def save_to_excel(sorted_schools, output_file, lista_escolas):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Adicione os cabeçalhos
    sheet["A1"] = "Nº"
    sheet["B1"] = "Nome da Escola"
    sheet["C1"] = "Número de Instalação"
    sheet["D1"] = "Valor da Fatura"
    sheet["E1"] = "Número da Conta"

    total = 0

    for index, (school, filepaths) in enumerate(sorted_schools.items(), start=2):
        school_total = 0
        found_pdf = False

        for filepath in filepaths:
            if filepath.lower().endswith('.pdf'):
                found_pdf = True

            invoice_value = extract_valor(filepath)
            if invoice_value is not None:
                school_total += invoice_value

            numero_instalacao = extract_numero_instalacao(filepath)
            numerodaconta = extract_numerodaconta(filepath)

        # Adicione as informações na planilha
        sheet[f"A{index}"] = lista_escolas[school][1]  # Número de faturas
        sheet[f"B{index}"] = lista_escolas[school][0]  # Nome da escola
        sheet[f"C{index}"] = numero_instalacao
        sheet[f"D{index}"] = school_total
        sheet[f"E{index}"] = numerodaconta

        if not found_pdf or len(filepaths) == 0:
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for col in ('A', 'B', 'C', 'D', 'E'):
                sheet[f'{col}{index}'].fill = yellow_fill

        total += school_total

    # Adicione o total geral na planilha
    sheet[f"C{index + 1}"] = "Total"
    last_row = len(sorted_schools) + 2
    sheet[f"D{last_row}"] = total

    # Salve a planilha no arquivo de saída
    wb.save(output_file)

# Atualize a chamada para save_to_excel
output_file = "C:\\path\\excel.xlsx"
save_to_excel(sorted_schools, output_file, lista_escolas)



if __name__ == "__main__":
    directory = "C:\\path"
    output_file = "C:\\path\\excel.xlsx"

    # Lendo o dicionário school_numbers do arquivo ou adicionando diretamente no código
    list_escolas = {
        # "104227": ("PROF. HENRIQUE GONÇALVES", 1),
        
}

    sorted_schools = get_sorted_schools_and_files(directory, lista_escolas)
   
    for school in lista_escolas:
        if school not in sorted_schools:
            sorted_schools[school] = []

    save_to_excel(sorted_schools, output_file, lista_escolas)